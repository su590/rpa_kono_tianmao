import dataclasses
import json
import os.path
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from requests import Session
from src.schedule import PATH
from src.spider.wph.WphGysLogin import WphGysSession


# 唯品会登录的账密实体类
@dataclasses.dataclass
class _WphAccount:
    username: str
    password: str
    port: int

@dataclasses.dataclass
class _CommodityInfo:
    merchandiseNo: str #MID
    goodsName: str
    prodSpuId: str #PSPUID

# 获取唯品会的商品排名
@dataclasses.dataclass
class _CommodityRank:
    rank: int
    imgUrl: str
    commodityInfo: _CommodityInfo
    brandStoreName: str
    minPayPrice: float
    pageMerDetailFlowIndex: int
    ctrIndex: int
    addUserNumIndex: int
    goodsActureAmtIndex: int
    goodsActureNumIndex: int
    convRateIndex: int
    unitPriceIndex: int

def _get_wph_account() -> _WphAccount:
    path = os.path.join(PATH, 'wph_account.json')
    with open(path, 'r', encoding='utf-8') as f:
        account_dict: dict = json.load(f)[0]
        return _WphAccount(**account_dict)

def _get_rank(session: Session, order_field: int, date_time: datetime) -> list:
    """
    在唯品会 - 魔方罗盘
    :param session:
    :param order_field: 人气榜单：1 销售榜单：2
    :return:
    """
    date_str: str = date_time.strftime('%Y%m%d')

    url = "https://compass.vip.com/industry/industrySituation/queryIndustryMerRanking"
    payload = json.dumps({
        "startDt": date_str,
        "endDt": date_str,
        "dtType": 0,
        "brandStoreSn": "10040918",
        "categoryLevel": 1,
        "categoryId": 7470,
        "orderField": order_field
    })
    headers = {
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'cache-control': 'no-cache',
        'content-type': 'application/json',
        'origin': 'https://compass.vip.com',
        'pragma': 'no-cache',
        'priority': 'u=1, i',
        'referer': 'https://compass.vip.com/frontend/index.html',
        'sec-ch-ua': '"Not)A;Brand";v="8", "Chromium";v="138", "Microsoft Edge";v="138"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'x-requested-with': 'XMLHttpRequest',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36 Edg/138.0.0.0',
    }

    response = session.request("POST", url, headers=headers, data=payload)
    response_json = response.json()
    return response_json['data']['rankingDataList']

def _assemble_data(rank_list: list) -> list:
    """
    返回组装的前100名数据 并保存在表格当中
    :param rank_list: 原始商品字典列表
    :return:
    """
    data_objects = []
    for item in rank_list[:100]:
        # 构造商品信息对象
        info = _CommodityInfo(merchandiseNo=item['merchandiseNo'], goodsName=item['goodsName'], prodSpuId=item['prodSpuId'])
        # 构造商品排名对象
        rank_obj = _CommodityRank(
            rank=item['ranking'],
            imgUrl=item['imgUrl'],
            commodityInfo=info,
            brandStoreName=item['brandStoreName'],
            minPayPrice=item['minPayPrice'],
            pageMerDetailFlowIndex=item['pageMerDetailFlowIndex'],
            ctrIndex=item['ctrIndex'],
            addUserNumIndex=item['addUserNumIndex'],
            goodsActureAmtIndex=item['goodsActureAmtIndex'],
            goodsActureNumIndex=item['goodsActureNumIndex'],
            convRateIndex=item['convRateIndex'],
            unitPriceIndex=item['unitPriceIndex']
        )

        data_objects.append(rank_obj)
    return data_objects

WPHPATH = "D:\ktm\wpx_rank_monthly_eight.xlsx"
def _save_to_excel(object_list: list, sheet_name: str, date: datetime, excel_path: str = WPHPATH) -> None:
    """
    把当天的排名信息追加保存到 Excel 中的指定 Sheet，并自动设置列宽
    :param object_list: _CommodityRank 对象列表
    :param sheet_name: Sheet 名称
    :param excel_path: Excel 文件路径
    """
    headers = [
        "日期", "排名", "商品编号", "商品名称", "prodSpuId", "图片地址", "品牌店铺",
        "最低支付价格", "详情页流量指数", "点击率指数", "加购人数指数", "成交金额指数",
        "成交件数指数", "转化率指数", "客单价指数"
    ]

    date_str = date.strftime('%Y-%m-%d')

    # 准备数据行
    rows = []
    for obj in object_list:
        rows.append([
            date_str,
            obj.rank,
            obj.commodityInfo.merchandiseNo,
            obj.commodityInfo.goodsName,
            obj.commodityInfo.prodSpuId,
            obj.imgUrl,
            obj.brandStoreName,
            obj.minPayPrice,
            obj.pageMerDetailFlowIndex,
            obj.ctrIndex,
            obj.addUserNumIndex,
            obj.goodsActureAmtIndex,
            obj.goodsActureNumIndex,
            obj.convRateIndex,
            obj.unitPriceIndex,
        ])

    # 打开或新建 Excel 文件
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        # 删除默认空白 sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

    # 获取或创建指定 Sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)  # 新建 Sheet 写入表头

    # 检查是否已写入今天的数据（防重复）
    existing_dates = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]
    if date_str in existing_dates:
        print(f"⚠️ {sheet_name} 已包含 {date_str} 的数据，跳过写入。")
        return

    # 写入新数据
    for row in rows:
        ws.append(row)

    # ✅ 设置自动列宽
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # 可加大一点留白
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(excel_path)
    print(f"✅ {sheet_name}：成功保存 {len(rows)} 条数据到 {excel_path}")


from tqdm import tqdm
import time
import random

if __name__ == '__main__':
    _wph_account = _get_wph_account()
    start_date = datetime(2025, 8, 1)
    end_date = datetime(2025, 8, 6)

    # date_range = list((start_date + timedelta(days=n)) for n in range((end_date - start_date).days + 1))
    date_range = [datetime(2025, 8, 8)]
    with WphGysSession(_wph_account.username, _wph_account.password, _wph_account.port) as session:
        for single_date in tqdm(date_range, desc="📅 日期进度", unit="天"):
            for order_field, sheet_name in tqdm([(1, "人气榜单"), (2, "销售榜单")], desc=f"📊 榜单进度（{single_date.strftime('%Y-%m-%d')}）", leave=False):
                try:
                    print(f"\n⏳ 获取 {single_date.strftime('%Y-%m-%d')} 的【{sheet_name}】...")
                    rank_data = _get_rank(session, order_field=order_field, date_time=single_date)
                    assembled = _assemble_data(rank_data)
                    _save_to_excel(assembled, sheet_name, date=single_date)
                    wait_time = round(random.uniform(2, 5), 2)
                    time.sleep(wait_time)
                except Exception as e:
                    print(f"❌ 获取或保存 {sheet_name} - {single_date.strftime('%Y-%m-%d')} 时出错: {e}")




